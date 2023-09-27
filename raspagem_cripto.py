from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys



class Cripto():
    def __init__(self) -> None:
        self.driver = webdriver.Chrome()
        self.site_link = ('https://www.kraken.com/prices?page=1')

    def abre(self):
        self.driver.get(self.site_link)
        self.driver.maximize_window()
        
    def main(self):
        self.abre()
        self.raspagem()
        self.cria_planilhas()
        self.envia_email()
        sleep(123123213)

    def raspagem(self):
        global armazena_nome, armazena_preco,armazena_change
        contador = 1
        paginas = 1
        armazena_nome = ['']
        armazena_preco = ['']
        armazena_change = ['']
        while True:
            self.site_map = {
                'xp':{
                    'nome': f'/html/body/div[1]/div/div/div/article/div/div[7]/div/div/div/div/div/div[3]/table/tbody/tr[{contador}]/td[2]/div/a/div/span[1]',
                    'preco': f'/html/body/div[1]/div/div/div/article/div/div[7]/div/div/div/div/div/div[3]/table/tbody/tr[{contador}]/td[3]/span',
                    'change': f'/html/body/div[1]/div/div/div/article/div/div[7]/div/div/div/div/div/div[3]/table/tbody/tr[{contador}]/td[5]/span',
                    'passa': '/html/body/div[1]/div/div/div/article/div/div[7]/div/div/div/div/div/div[4]/nav/ul/li[18]',
                }               
            }
            try:
                nome = self.driver.find_element(By.XPATH,self.site_map['xp']['nome']).text
                preco = self.driver.find_element(By.XPATH,self.site_map['xp']['preco']).text
                change = self.driver.find_element(By.XPATH,self.site_map['xp']['change']).text

                armazena_nome.append(nome)
                armazena_preco.append(preco)
                armazena_change.append(change)

                print(nome)
                print(preco)
                print(change)
                contador += 1
                sleep(0.5)

            except NoSuchElementException:
                self.lazy_upload()
                sleep(2)
                try:
                    passa = self.driver.find_element(By.XPATH, self.site_map['xp']['passa'])
                    if paginas >= 23:
                        print('pagina exedeu o numero limite')
                        break

                    if passa:
                        passa.click()
                        sleep(3)
                        contador = 1
                        paginas += 1
                        print(f'voce esta na pagina {paginas}')
                    else:
                        break
                except:
                    print('xpath nao encontrado!!! ')
                    
                

            except:
                print('houve algo inesperado!!')
                break


    def lazy_upload(self):
        for i in range(0,3):
            actions = ActionChains(self.driver)
            actions.send_keys(Keys.PAGE_DOWN).perform()

    def cria_planilhas(self):
        planilha = openpyxl.Workbook()
        cripto = planilha.active
        cripto.title = 'cripto'
        cripto['A1'] = 'Nome'
        cripto['B1'] = 'Preco'
        cripto['c1'] = 'change'

        for index, (nome, preco, change) in enumerate(zip(armazena_nome, armazena_preco, armazena_change), start=2):
            cripto.cell(column=1, row=index, value=nome)
            cripto.cell(column=2, row=index, value=preco)
            cripto.cell(column=3, row=index, value=change)

        planilha.save('planilha_de_preco.xlsx')
        print('Planilha criada com sucesso!')

    def envia_email(self):
        smtp_server = 'smtp.gmail.com'
        smtp_port = 587
        email_sender =input('digite o seu gmail: ')
        email_password =('ufiwahyqoyzgaihn')
        recipient_email = input('pra que voce quer enviar o email?: ') 
        email_subject = 'planilha de dados'


        email_body = 'segue em anaxo a planilha de dados: '


        msg = MIMEMultipart()
        msg['From'] = email_sender
        msg['To'] = recipient_email
        msg['Subject'] = email_subject
        msg.attach(MIMEText(email_body, 'plain'))

        filename = 'planilha_de_preco.xlsx'
        attachment = open(filename, 'rb')

        part = MIMEBase('application', 'octet-stream')
        part.set_payload((attachment).read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', "attachment; filename= %s" % filename)

        msg.attach(part)

        try:

            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            
            server.login(email_sender, email_password)
            
            server.sendmail(email_sender, recipient_email, msg.as_string())
            
            server.quit()
            
            print('E-mail com anexo enviado com sucesso!')
        except Exception as e:
            print('Erro ao enviar o e-mail:', str(e))


cripto = Cripto()
cripto.main()
